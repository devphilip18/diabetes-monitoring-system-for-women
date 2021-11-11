from flask import Flask, render_template, request

import pandas as pd
from openpyxl import load_workbook
import sklearn.linear_model 
from ObjectDetector import Detector
import numpy as np
from PIL import Image

app = Flask(__name__)

detector = Detector()


@app.route('/', methods=['GET'])
def index():
	if(request.method == 'GET'):
		return render_template('index.html')
		

@app.route('/predict', methods=['GET'])
def user_input():
	if(request.method == 'GET'):
		return render_template('user_input.html')
		
		
@app.route('/detect', methods=['GET'])
def detect():
	if(request.method == 'GET'):
		return render_template('final.html')


@app.route('/getdata', methods=['GET', 'POST'])
def check():
	foods = ['Burger', 'Beer', 'Pizza', 'Coffee','Apple','Potatoes','Pastry','Pasta','Mango','Chips','Salad']
	preg = request.form.get('preganency')
	bloodP =request.form.get('BloodPressure')
	skinT = request.form.get('skinthickness')
	insulin = request.form.get('insulin')
	bmi = request.form.get('bmi')
	age = request.form.get('age')
	predict_list = [int(preg),int(bloodP),int(skinT), int(insulin), int(bmi),int(age)]
	var = precitMe(predict_list)
	print("Var",var);
	appendDataToExcel(preg,bloodP,skinT, insulin, bmi,'0.0', age,var)

	msg = ''
	if var < 90:
		msg = "Your blood sugar is low. Increase your sugar intake."
	elif var >= 90 and var <= 150:
		msg = "Your sugar level is normal! Keep up the good work."
	elif var > 150 and var <= 240:
		msg = "Your blood sugar is too high. Work on lowering down your sugar intake."
	else:
		msg = "Your sugar level is out of control."
	

	return render_template('display_data.html', msg = msg,foods=foods,gluc=var)
	
	

def appendDataToExcel(preg,bloodP,skinT, insulin, bmi,pedi, age,gluc):
	
	file = 'E:\\path\\to\\diabetes_data.xlsx' #TODO: update the path to diabetes dataset
	new_row = [preg,bloodP,skinT, insulin,bmi,pedi,age,gluc]
	wb = openpyxl.load_workbook(filename=file)
	ws = wb['diabetes']
	row = ws.max_row +1
	for col, entry in enumerate(new_row, start=1):
		ws.cell(row=row, column=col, value=entry)
	wb.save(file)



def precitMe(predict_list):
	value = predictDiabetes(predict_list)
	var = int(value[0])
	wb = load_workbook(filename='E:\\path\\to\\diabetes_data.xlsx') #TODO: update the path to diabetes dataset
	ws = wb.worksheets[0]
	row = ws.max_row
	ws.cell(row=row, column=7).value = var

	return var



@app.route('/foodItem/detect', methods=['POST'])
def foodDetect():
	if request.method == 'POST':
		file = Image.open(request.files['file'].stream)
		img,label = detector.detectObject(file)
		print(label)
		
		res=label.split(':')
		res=res[0].strip()
		print(res)
		
		file2 = 'E:\\DevPhilip\\test\\food.xlsx' #TODO: update the path to food dataset
    
		wb2 = openpyxl.load_workbook(filename=file2)
		ws2 = wb2['Sheet1']

		calorie = 0

		for i in range(1, ws2.max_row+1):
			if ws2.cell(row=i, column=1).value == res:
				calorie = ws2.cell(row=i, column=4).value
				print("calorie=",calorie)
		
	
		return render_template('final.html', calo = calorie,food = label,type="food_detect")
		
		

	
	
def predictDiabetes(predict_list):
	to_predict = np.array(predict_list).reshape(1, 6) 

	col_names = ['Pregnancies', 'BloodPressure', 'SkinThickness', 'Insulin', 'BMI', 'DiabetesPedigreeFunction', 'Age','Glucose']
	pima = pd.read_excel('E:\\path\\to\\diabetes_data.xlsx', names=col_names, engine='openpyxl') #TODO: update the path to diabetes dataset
	per = 0.55
	plot = pd.DataFrame(col_names)
	
	zero_not_accepted = [ 'BloodPressure','SkinThickness', 'BMI', 'Insulin','Glucose']


	for column in zero_not_accepted:
		pima[column] = pima[column].replace(0, np.NaN)
		mean = int(pima[column].mean(skipna=True))
		pima[column] = pima[column].replace(np.NaN, mean)
		
	train = pima[:((int)(len(plot) * per))]
	xtrain = train[['Pregnancies', 'BloodPressure', 'SkinThickness', 'Insulin', 'BMI', 'Age']]
	ytrain = train[['Glucose']]

	test = pima[((int)(len(plot) * per)):]
	xtest = test[['Pregnancies', 'BloodPressure', 'SkinThickness', 'Insulin', 'BMI', 'Age']]
	ytest = test[['Glucose']]
	dataframeHeading = plot.head()

	ols = sklearn.linear_model.LinearRegression()
	model = ols.fit(xtrain, ytrain)
	
	print(to_predict)
	predict = model.predict(to_predict)

	
	print("predict",predict[0])
	
	
	return predict

if '__main__' == __name__:
	app.run(debug=True)